from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date, timedelta
from io import BytesIO
from decimal import Decimal
import jdatetime
from utils.decimal_utils import to_decimal

class ExcelReportGenerator:
    
    def __init__(self, pareto_service, abc_service):
        self.pareto_service = pareto_service
        self.abc_service = abc_service
    
    def generate_pareto_report(self, mode='خرید', category='Food', days=30):
        """
        Generate complete Excel report with:
        - Dashboard sheet with KPIs (both Food & NonFood)
        - Pareto Analysis Food with chart
        - Pareto Analysis NonFood with chart
        - ABC Classification Food with colors
        - ABC Classification NonFood with colors
        - Comparison sheet
        - Flowchart sheet
        - Raw data sheets
        """
        wb = Workbook()
        
        # Dashboard with both categories
        self._create_dashboard_sheet(wb, mode, days)
        
        # Pareto sheets for both categories
        self._create_pareto_sheet(wb, mode, 'Food', days)
        self._create_pareto_sheet(wb, mode, 'NonFood', days)
        
        # ABC sheets for both categories
        self._create_abc_sheet(wb, mode, 'Food', days)
        self._create_abc_sheet(wb, mode, 'NonFood', days)
        
        # Comparison and analysis
        self._create_comparison_sheet(wb, days)
        self._create_flowchart_sheet(wb)
        
        # Raw data for both categories
        self._create_data_sheet(wb, mode, 'Food', days)
        self._create_data_sheet(wb, mode, 'NonFood', days)
        
        for sheet in wb.worksheets:
            sheet.sheet_view.rightToLeft = True
        
        return wb
    
    def _create_dashboard_sheet(self, wb, mode, days):
        ws = wb.active
        ws.title = "Dashboard"
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:L1')
        ws['A1'] = '🏨 گزارش جامع تحلیل موجودی هتل'
        ws['A1'].font = Font(name='B Nazanin', size=20, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        persian_date = jdatetime.date.today().strftime('%Y/%m/%d')
        ws.merge_cells('A2:L2')
        ws['A2'] = f'تاریخ گزارش: {persian_date} | دوره: {days} روز | نوع: {mode} | شامل: غذایی و غیرغذایی'
        ws['A2'].font = Font(name='B Nazanin', size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 25
        
        # Get stats for both categories
        food_stats = self.pareto_service.get_summary_stats(mode, 'Food', days)
        nonfood_stats = self.pareto_service.get_summary_stats(mode, 'NonFood', days)
        
        # === FOOD Section ===
        ws.merge_cells('A4:F4')
        ws['A4'] = '🍽️ اقلام غذایی (Food)'
        ws['A4'].font = Font(name='B Nazanin', size=14, bold=True, color='FFFFFF')
        ws['A4'].fill = PatternFill(start_color='2E7D32', fill_type='solid')
        ws['A4'].alignment = Alignment(horizontal='center')
        
        food_kpis = [
            ('A6', 'B6', 'تعداد اقلام', food_stats['total_items'], 'E8F5E9', '2E7D32'),
            ('C6', 'D6', 'مجموع مبلغ', f"{food_stats['total_amount']:,.0f}", 'E3F2FD', '1565C0'),
            ('A8', 'B8', 'کلاس A', f"{food_stats['class_a_count']} قلم", 'C8E6C9', '1B5E20'),
            ('C8', 'D8', 'کلاس B', f"{food_stats['class_b_count']} قلم", 'FFF9C4', 'F57F17'),
            ('E8', 'F8', 'کلاس C', f"{food_stats['class_c_count']} قلم", 'FFCCBC', 'BF360C'),
        ]
        
        for title_cell, value_cell, title, value, bg_color, text_color in food_kpis:
            ws[title_cell] = title
            ws[title_cell].font = Font(name='B Nazanin', size=10, bold=True)
            ws[title_cell].fill = PatternFill(start_color=bg_color, fill_type='solid')
            ws[title_cell].border = thin_border
            ws[title_cell].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[value_cell] = value
            ws[value_cell].font = Font(name='B Nazanin', size=12, bold=True, color=text_color)
            ws[value_cell].fill = PatternFill(start_color=bg_color, fill_type='solid')
            ws[value_cell].border = thin_border
            ws[value_cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # === NONFOOD Section ===
        ws.merge_cells('H4:L4')
        ws['H4'] = '🧴 اقلام غیرغذایی (NonFood)'
        ws['H4'].font = Font(name='B Nazanin', size=14, bold=True, color='FFFFFF')
        ws['H4'].fill = PatternFill(start_color='1565C0', fill_type='solid')
        ws['H4'].alignment = Alignment(horizontal='center')
        
        nonfood_kpis = [
            ('H6', 'I6', 'تعداد اقلام', nonfood_stats['total_items'], 'E3F2FD', '1565C0'),
            ('J6', 'K6', 'مجموع مبلغ', f"{nonfood_stats['total_amount']:,.0f}", 'E3F2FD', '1565C0'),
            ('H8', 'I8', 'کلاس A', f"{nonfood_stats['class_a_count']} قلم", 'C8E6C9', '1B5E20'),
            ('J8', 'K8', 'کلاس B', f"{nonfood_stats['class_b_count']} قلم", 'FFF9C4', 'F57F17'),
            ('L8', 'L8', 'کلاس C', f"{nonfood_stats['class_c_count']} قلم", 'FFCCBC', 'BF360C'),
        ]
        
        for title_cell, value_cell, title, value, bg_color, text_color in nonfood_kpis:
            ws[title_cell] = title
            ws[title_cell].font = Font(name='B Nazanin', size=10, bold=True)
            ws[title_cell].fill = PatternFill(start_color=bg_color, fill_type='solid')
            ws[title_cell].border = thin_border
            ws[title_cell].alignment = Alignment(horizontal='center', vertical='center')
            
            if title_cell != value_cell:
                ws[value_cell] = value
                ws[value_cell].font = Font(name='B Nazanin', size=12, bold=True, color=text_color)
                ws[value_cell].fill = PatternFill(start_color=bg_color, fill_type='solid')
                ws[value_cell].border = thin_border
                ws[value_cell].alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col].width = 14
        
        for row in [4, 6, 8]:
            ws.row_dimensions[row].height = 35
        
        # === Pie Charts ===
        # Food ABC Pie
        ws['A11'] = 'کلاس'
        ws['B11'] = 'غذایی'
        ws['A12'] = 'A'
        ws['A13'] = 'B'
        ws['A14'] = 'C'
        ws['B12'] = food_stats.get('class_a_amount', 0)
        ws['B13'] = food_stats.get('class_b_amount', 0)
        ws['B14'] = food_stats.get('class_c_amount', 0)
        
        pie_food = PieChart()
        pie_food.title = "توزیع ABC غذایی"
        pie_food.width = 10
        pie_food.height = 8
        pie_food.dataLabels = DataLabelList(showPercent=True, showCatName=True)
        pie_food.add_data(Reference(ws, min_col=2, min_row=11, max_row=14), titles_from_data=True)
        pie_food.set_categories(Reference(ws, min_col=1, min_row=12, max_row=14))
        ws.add_chart(pie_food, "A16")
        
        # NonFood ABC Pie
        ws['D11'] = 'کلاس'
        ws['E11'] = 'غیرغذایی'
        ws['D12'] = 'A'
        ws['D13'] = 'B'
        ws['D14'] = 'C'
        ws['E12'] = nonfood_stats.get('class_a_amount', 0)
        ws['E13'] = nonfood_stats.get('class_b_amount', 0)
        ws['E14'] = nonfood_stats.get('class_c_amount', 0)
        
        pie_nonfood = PieChart()
        pie_nonfood.title = "توزیع ABC غیرغذایی"
        pie_nonfood.width = 10
        pie_nonfood.height = 8
        pie_nonfood.dataLabels = DataLabelList(showPercent=True, showCatName=True)
        pie_nonfood.add_data(Reference(ws, min_col=5, min_row=11, max_row=14), titles_from_data=True)
        pie_nonfood.set_categories(Reference(ws, min_col=4, min_row=12, max_row=14))
        ws.add_chart(pie_nonfood, "G16")
    
    def _create_pareto_sheet(self, wb, mode, category, days):
        ws = wb.create_sheet(f"Pareto {category}")
        
        df = self.pareto_service.calculate_pareto(mode, category, days)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:H1')
        ws['A1'] = f'تحلیل پارتو: {mode} - {category} (آخرین {days} روز)'
        ws['A1'].font = Font(name='B Nazanin', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2E7D32', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        headers = ['ردیف', 'کد کالا', 'نام کالا', 'مبلغ', 'درصد سهم', 
                   'مبلغ تجمعی', 'درصد تجمعی', 'کلاس ABC']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(name='B Nazanin', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='455A64', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[3].height = 28
        
        if not df.empty:
            for r_idx, row in enumerate(df.itertuples(), 4):
                ws.cell(row=r_idx, column=1, value=row.row_num).border = thin_border
                ws.cell(row=r_idx, column=2, value=row.item_code).border = thin_border
                ws.cell(row=r_idx, column=3, value=row.item_name).border = thin_border
                
                amount_cell = ws.cell(row=r_idx, column=4, value=row.amount)
                amount_cell.number_format = '#,##0'
                amount_cell.border = thin_border
                
                pct_cell = ws.cell(row=r_idx, column=5, value=row.percentage / 100)
                pct_cell.number_format = '0.00%'
                pct_cell.border = thin_border
                
                cum_amount_cell = ws.cell(row=r_idx, column=6, value=row.cumulative_amount)
                cum_amount_cell.number_format = '#,##0'
                cum_amount_cell.border = thin_border
                
                cum_pct_cell = ws.cell(row=r_idx, column=7, value=row.cumulative_percentage / 100)
                cum_pct_cell.number_format = '0.00%'
                cum_pct_cell.border = thin_border
                
                abc_cell = ws.cell(row=r_idx, column=8, value=row.abc_class)
                abc_cell.border = thin_border
                abc_cell.alignment = Alignment(horizontal='center')
                
                if row.abc_class == 'A':
                    abc_cell.fill = PatternFill(start_color='C8E6C9', fill_type='solid')
                    abc_cell.font = Font(bold=True, color='1B5E20')
                elif row.abc_class == 'B':
                    abc_cell.fill = PatternFill(start_color='FFF9C4', fill_type='solid')
                    abc_cell.font = Font(bold=True, color='F57F17')
                else:
                    abc_cell.fill = PatternFill(start_color='FFCCBC', fill_type='solid')
                    abc_cell.font = Font(color='BF360C')
                
                for col in range(1, 9):
                    ws.cell(row=r_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            if len(df) >= 2:
                # Create combo chart: Bar for amounts + Line for cumulative percentage
                
                # 1. Bar Chart for amounts
                bar_chart = BarChart()
                bar_chart.type = "col"
                bar_chart.style = 10
                bar_chart.title = f"نمودار پارتو - {category}"
                bar_chart.y_axis.title = 'مبلغ (ریال)'
                bar_chart.x_axis.title = 'کالا'
                
                data = Reference(ws, min_col=4, min_row=3, max_row=3 + len(df), max_col=4)
                cats = Reference(ws, min_col=3, min_row=4, max_row=3 + len(df))
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(cats)
                bar_chart.shape = 4
                bar_chart.width = 20
                bar_chart.height = 12
                
                # 2. Line Chart for cumulative percentage (secondary axis)
                line_chart = LineChart()
                line_chart.style = 10
                line_chart.y_axis.axId = 200
                line_chart.y_axis.title = 'درصد تجمعی'
                
                cum_data = Reference(ws, min_col=7, min_row=3, max_row=3 + len(df), max_col=7)
                line_chart.add_data(cum_data, titles_from_data=True)
                line_chart.set_categories(cats)
                
                # Style line
                s = line_chart.series[0]
                s.graphicalProperties.line.width = 25000  # 2.5pt
                s.marker.symbol = "circle"
                s.marker.size = 7
                
                # Combine charts
                bar_chart.y_axis.crosses = "min"
                line_chart.y_axis.crosses = "max"
                bar_chart += line_chart
                
                ws.add_chart(bar_chart, "J5")
                
                # 3. Pie Chart for ABC distribution
                pie_chart = PieChart()
                pie_chart.title = "توزیع کلاس‌های ABC"
                pie_chart.width = 12
                pie_chart.height = 10
                
                # Calculate ABC totals for pie chart
                abc_row = 3 + len(df) + 3
                ws.cell(row=abc_row, column=10, value="کلاس A")
                ws.cell(row=abc_row, column=11, value=df[df['abc_class'] == 'A']['amount'].sum())
                ws.cell(row=abc_row + 1, column=10, value="کلاس B")
                ws.cell(row=abc_row + 1, column=11, value=df[df['abc_class'] == 'B']['amount'].sum())
                ws.cell(row=abc_row + 2, column=10, value="کلاس C")
                ws.cell(row=abc_row + 2, column=11, value=df[df['abc_class'] == 'C']['amount'].sum())
                
                pie_labels = Reference(ws, min_col=10, min_row=abc_row, max_row=abc_row + 2)
                pie_data = Reference(ws, min_col=11, min_row=abc_row - 1, max_row=abc_row + 2)
                pie_chart.add_data(pie_data, titles_from_data=True)
                pie_chart.set_categories(pie_labels)
                
                # Add data labels
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.dataLabels.showCatName = True
                
                ws.add_chart(pie_chart, "J20")
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 12
    
    def _create_abc_sheet(self, wb, mode, category, days):
        ws = wb.create_sheet(f"ABC {category}")
        
        classified = self.abc_service.get_abc_classification(mode, category, days)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:G1')
        ws['A1'] = f'طبقه‌بندی ABC: {mode} - {category}'
        ws['A1'].font = Font(name='B Nazanin', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='6A1B9A', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        current_row = 3
        
        class_configs = [
            ('A', '✅ کلاس A: اقلام حیاتی (80% ارزش)', 'C8E6C9', '1B5E20', 
             'توصیه: کنترل روزانه، سفارش دقیق، تأمین‌کننده بکاپ'),
            ('B', '⚠️ کلاس B: اقلام مهم (15% ارزش)', 'FFF9C4', 'F57F17',
             'توصیه: کنترل هفتگی، سفارش معمولی'),
            ('C', '⚪ کلاس C: اقلام معمولی (5% ارزش)', 'FFCCBC', 'BF360C',
             'توصیه: کنترل ماهانه، سفارش انبوه')
        ]
        
        for abc_class, title, bg_color, text_color, recommendation in class_configs:
            ws.merge_cells(f'A{current_row}:G{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = title
            title_cell.font = Font(name='B Nazanin', size=14, bold=True, color=text_color)
            title_cell.fill = PatternFill(start_color=bg_color, fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = thin_border
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:G{current_row}')
            rec_cell = ws[f'A{current_row}']
            rec_cell.value = recommendation
            rec_cell.font = Font(name='B Nazanin', size=10, italic=True)
            rec_cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            headers = ['کد کالا', 'نام کالا', 'واحد', 'مقدار', 'مبلغ کل', 'درصد', 'تجمعی']
            header_row = current_row
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = Font(name='B Nazanin', size=10, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='546E7A', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            current_row += 1
            
            items = classified.get(abc_class, [])
            data_start_row = header_row + 1
            for item in items:
                ws.cell(row=current_row, column=1, value=item['item_code']).border = thin_border
                ws.cell(row=current_row, column=2, value=item['item_name']).border = thin_border
                ws.cell(row=current_row, column=3, value=item['unit']).border = thin_border
                
                qty_cell = ws.cell(row=current_row, column=4, value=item['total_quantity'])
                qty_cell.number_format = '#,##0.00'
                qty_cell.border = thin_border
                
                amount_cell = ws.cell(row=current_row, column=5, value=item['total_amount'])
                amount_cell.number_format = '#,##0'
                amount_cell.border = thin_border
                
                pct_cell = ws.cell(row=current_row, column=6, value=item['percentage'] / 100)
                pct_cell.number_format = '0.00%'
                pct_cell.border = thin_border
                
                cum_cell = ws.cell(row=current_row, column=7, value=item['cumulative_percentage'] / 100)
                cum_cell.number_format = '0.00%'
                cum_cell.border = thin_border
                
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                
                current_row += 1
            
            # Chart: Top items within the class (by amount)
            top_n = min(5, len(items))
            if top_n > 0:
                data_end_row = data_start_row + top_n - 1
                data_ref = Reference(ws, min_col=5, min_row=header_row, max_row=data_start_row + top_n - 1)
                cats_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                bar_chart = BarChart()
                bar_chart.title = f"Top {top_n} کلاس {abc_class}"
                bar_chart.y_axis.title = 'مبلغ'
                bar_chart.x_axis.title = 'کالا'
                bar_chart.width = 14
                bar_chart.height = 8
                bar_chart.add_data(data_ref, titles_from_data=True)
                bar_chart.set_categories(cats_ref)
                chart_anchor_row = header_row
                ws.add_chart(bar_chart, f"J{chart_anchor_row}")
            
            if not items:
                ws.merge_cells(f'A{current_row}:G{current_row}')
                ws[f'A{current_row}'] = 'هیچ کالایی در این کلاس وجود ندارد'
                ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
                ws[f'A{current_row}'].font = Font(italic=True, color='757575')
                current_row += 1
            
            current_row += 1
        
        for col, width in [('A', 12), ('B', 20), ('C', 10), ('D', 12), ('E', 15), ('F', 10), ('G', 10)]:
            ws.column_dimensions[col].width = width
    
    def _create_comparison_sheet(self, wb, days):
        """Create a sheet with multiple comparison charts for multi-dimensional analysis"""
        ws = wb.create_sheet("Comparison")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:U1')
        ws['A1'] = '📊 تحلیل چندجانبه و مقایسه‌ای'
        ws['A1'].font = Font(name='B Nazanin', size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='0D47A1', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # === Section 1: Food vs NonFood comparison ===
        ws.merge_cells('A3:I3')
        ws['A3'] = '🍽️ مقایسه غذایی و غیرغذایی (خرید)'
        ws['A3'].font = Font(name='B Nazanin', size=14, bold=True, color='FFFFFF')
        ws['A3'].fill = PatternFill(start_color='1976D2', fill_type='solid')
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Get data for both categories
        food_stats = self.pareto_service.get_summary_stats('خرید', 'Food', days)
        nonfood_stats = self.pareto_service.get_summary_stats('خرید', 'NonFood', days)
        
        # Data table for Food vs NonFood
        headers = ['دسته‌بندی', 'تعداد اقلام', 'مجموع مبلغ', 'کلاس A', 'کلاس B', 'کلاس C']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = Font(name='B Nazanin', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='455A64', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Food row
        food_data = ['غذایی', food_stats['total_items'], food_stats['total_amount'],
                     food_stats['class_a_count'], food_stats['class_b_count'], food_stats['class_c_count']]
        for col, val in enumerate(food_data, 1):
            cell = ws.cell(row=6, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if col == 3:
                cell.number_format = '#,##0'
        
        # NonFood row
        nonfood_data = ['غیرغذایی', nonfood_stats['total_items'], nonfood_stats['total_amount'],
                        nonfood_stats['class_a_count'], nonfood_stats['class_b_count'], nonfood_stats['class_c_count']]
        for col, val in enumerate(nonfood_data, 1):
            cell = ws.cell(row=7, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if col == 3:
                cell.number_format = '#,##0'
        
        # Chart 1: Total Amount comparison (Food vs NonFood)
        bar1 = BarChart()
        bar1.type = "col"
        bar1.title = "مقایسه مبلغ کل: غذایی vs غیرغذایی"
        bar1.y_axis.title = 'مبلغ (ریال)'
        bar1.width = 12
        bar1.height = 10
        data1 = Reference(ws, min_col=3, min_row=5, max_row=7)
        cats1 = Reference(ws, min_col=1, min_row=6, max_row=7)
        bar1.add_data(data1, titles_from_data=True)
        bar1.set_categories(cats1)
        ws.add_chart(bar1, "J5")
        
        # Chart 2: ABC class count comparison
        bar2 = BarChart()
        bar2.type = "col"
        bar2.style = 10
        bar2.title = "مقایسه تعداد اقلام در کلاس‌های ABC"
        bar2.y_axis.title = 'تعداد'
        bar2.width = 14
        bar2.height = 10
        data2 = Reference(ws, min_col=4, min_row=5, max_row=7, max_col=6)
        bar2.add_data(data2, titles_from_data=True)
        bar2.set_categories(cats1)
        ws.add_chart(bar2, "R5")
        
        # === Section 2: Transaction Type Comparison ===
        ws.merge_cells('A25:I25')
        ws['A25'] = '📈 مقایسه انواع تراکنش'
        ws['A25'].font = Font(name='B Nazanin', size=14, bold=True, color='FFFFFF')
        ws['A25'].fill = PatternFill(start_color='388E3C', fill_type='solid')
        ws['A25'].alignment = Alignment(horizontal='center')
        
        # Get stats for different transaction types
        modes = ['خرید', 'مصرف', 'ضایعات']
        mode_stats = {}
        for m in modes:
            try:
                mode_stats[m] = self.pareto_service.get_summary_stats(m, 'Food', days)
            except Exception:
                mode_stats[m] = {'total_items': 0, 'total_amount': 0, 'class_a_count': 0, 'class_b_count': 0, 'class_c_count': 0}
        
        # Data table for transaction types
        headers2 = ['نوع تراکنش', 'تعداد اقلام', 'مجموع مبلغ', 'کلاس A', 'کلاس B', 'کلاس C']
        for col, h in enumerate(headers2, 1):
            cell = ws.cell(row=27, column=col, value=h)
            cell.font = Font(name='B Nazanin', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='455A64', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for r_idx, m in enumerate(modes, 28):
            stats = mode_stats[m]
            row_data = [m, stats['total_items'], stats['total_amount'],
                        stats['class_a_count'], stats['class_b_count'], stats['class_c_count']]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col == 3:
                    cell.number_format = '#,##0'
        
        # Chart 3: Transaction type amounts
        bar3 = BarChart()
        bar3.type = "col"
        bar3.title = "مقایسه مبلغ بر اساس نوع تراکنش"
        bar3.y_axis.title = 'مبلغ (ریال)'
        bar3.width = 12
        bar3.height = 10
        data3 = Reference(ws, min_col=3, min_row=27, max_row=30)
        cats3 = Reference(ws, min_col=1, min_row=28, max_row=30)
        bar3.add_data(data3, titles_from_data=True)
        bar3.set_categories(cats3)
        ws.add_chart(bar3, "J27")
        
        # Chart 4: Pie chart for transaction type distribution
        pie1 = PieChart()
        pie1.title = "توزیع مبلغ بر اساس نوع تراکنش"
        pie1.width = 12
        pie1.height = 10
        pie1.dataLabels = DataLabelList(showPercent=True, showCatName=True)
        pie1.add_data(data3, titles_from_data=True)
        pie1.set_categories(cats3)
        ws.add_chart(pie1, "R27")
        
        # === Section 3: Top Items Across Categories ===
        ws.merge_cells('A48:I48')
        ws['A48'] = '🏆 برترین اقلام در هر دسته'
        ws['A48'].font = Font(name='B Nazanin', size=14, bold=True, color='FFFFFF')
        ws['A48'].fill = PatternFill(start_color='7B1FA2', fill_type='solid')
        ws['A48'].alignment = Alignment(horizontal='center')
        
        # Get top 5 from each category
        food_df = self.pareto_service.calculate_pareto('خرید', 'Food', days)
        nonfood_df = self.pareto_service.calculate_pareto('خرید', 'NonFood', days)
        
        # Food Top 5
        ws.cell(row=50, column=1, value='غذایی - Top 5').font = Font(bold=True, color='1B5E20')
        ws.cell(row=50, column=1).fill = PatternFill(start_color='C8E6C9', fill_type='solid')
        for col, h in enumerate(['کد', 'نام', 'مبلغ'], 2):
            cell = ws.cell(row=50, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='C8E6C9', fill_type='solid')
            cell.border = thin_border
        
        if not food_df.empty:
            for i, row in enumerate(food_df.head(5).itertuples(), 51):
                ws.cell(row=i, column=2, value=row.item_code).border = thin_border
                ws.cell(row=i, column=3, value=row.item_name).border = thin_border
                c = ws.cell(row=i, column=4, value=row.amount)
                c.number_format = '#,##0'
                c.border = thin_border
        
        # NonFood Top 5
        ws.cell(row=50, column=6, value='غیرغذایی - Top 5').font = Font(bold=True, color='BF360C')
        ws.cell(row=50, column=6).fill = PatternFill(start_color='FFCCBC', fill_type='solid')
        for col, h in enumerate(['کد', 'نام', 'مبلغ'], 7):
            cell = ws.cell(row=50, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFCCBC', fill_type='solid')
            cell.border = thin_border
        
        if not nonfood_df.empty:
            for i, row in enumerate(nonfood_df.head(5).itertuples(), 51):
                ws.cell(row=i, column=7, value=row.item_code).border = thin_border
                ws.cell(row=i, column=8, value=row.item_name).border = thin_border
                c = ws.cell(row=i, column=9, value=row.amount)
                c.number_format = '#,##0'
                c.border = thin_border
        
        # Chart 5: Side-by-side top items comparison
        if not food_df.empty and len(food_df) >= 2:
            bar4 = BarChart()
            bar4.type = "bar"
            bar4.title = "Top 5 غذایی"
            bar4.y_axis.title = 'کالا'
            bar4.x_axis.title = 'مبلغ'
            bar4.width = 10
            bar4.height = 8
            data4 = Reference(ws, min_col=4, min_row=50, max_row=min(55, 50 + len(food_df)))
            cats4 = Reference(ws, min_col=3, min_row=51, max_row=min(55, 50 + len(food_df)))
            bar4.add_data(data4, titles_from_data=True)
            bar4.set_categories(cats4)
            ws.add_chart(bar4, "J50")
        
        if not nonfood_df.empty and len(nonfood_df) >= 2:
            bar5 = BarChart()
            bar5.type = "bar"
            bar5.title = "Top 5 غیرغذایی"
            bar5.y_axis.title = 'کالا'
            bar5.x_axis.title = 'مبلغ'
            bar5.width = 10
            bar5.height = 8
            data5 = Reference(ws, min_col=9, min_row=50, max_row=min(55, 50 + len(nonfood_df)))
            cats5 = Reference(ws, min_col=8, min_row=51, max_row=min(55, 50 + len(nonfood_df)))
            bar5.add_data(data5, titles_from_data=True)
            bar5.set_categories(cats5)
            ws.add_chart(bar5, "R50")
        
        # Column widths
        for col, w in [
            ('A', 14), ('B', 12), ('C', 15), ('D', 10), ('E', 10), ('F', 14), ('G', 12), ('H', 18), ('I', 15),
            ('J', 12), ('K', 12), ('L', 12), ('M', 12), ('N', 12), ('O', 12), ('P', 12), ('Q', 12), ('R', 12), ('S', 12), ('T', 12), ('U', 12)
        ]:
            ws.column_dimensions[col].width = w
    
    def _create_flowchart_sheet(self, wb):
        ws = wb.create_sheet("Flowchart")
        
        thin_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        ws.merge_cells('A1:J1')
        ws['A1'] = '🔄 فرآیند تصمیم‌گیری بر اساس کلاس ABC'
        ws['A1'].font = Font(name='B Nazanin', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='6A1B9A', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        flowchart_items = [
            (4, 'D', 'E', '🎯 شروع: شناسایی کالا', 'FFE082'),
            (6, 'D', 'E', '📊 محاسبه پارتو و ABC', 'B2DFDB'),
            (8, 'D', 'E', 'کلاس کالا چیست؟', 'FFE0B2'),
            (10, 'B', 'C', '✅ کلاس A\n(80% ارزش)', 'C8E6C9'),
            (10, 'E', 'F', '⚠️ کلاس B\n(15% ارزش)', 'FFF9C4'),
            (10, 'H', 'I', '⚪ کلاس C\n(5% ارزش)', 'FFCCBC'),
            (12, 'B', 'C', 'کنترل روزانه\nسفارش دقیق', 'A5D6A7'),
            (12, 'E', 'F', 'کنترل هفتگی\nسفارش معمولی', 'FFF59D'),
            (12, 'H', 'I', 'کنترل ماهانه\nسفارش انبوه', 'FFAB91'),
        ]
        
        for row, start_col, end_col, text, color in flowchart_items:
            ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
            cell = ws[f'{start_col}{row}']
            cell.value = text
            cell.font = Font(name='B Nazanin', size=11, bold=True)
            cell.fill = PatternFill(start_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            ws.row_dimensions[row].height = 40
        
        arrows = [
            (5, 'D', '⬇'),
            (7, 'D', '⬇'),
            (9, 'B', '↙'),
            (9, 'E', '⬇'),
            (9, 'H', '↘'),
            (11, 'B', '⬇'),
            (11, 'E', '⬇'),
            (11, 'H', '⬇'),
        ]
        
        for row, col, arrow in arrows:
            cell = ws[f'{col}{row}']
            cell.value = arrow
            cell.font = Font(size=18, color='1976D2')
            cell.alignment = Alignment(horizontal='center')
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 12
    
    def _create_data_sheet(self, wb, mode, category, days):
        ws = wb.create_sheet(f"Raw Data {category}")
        
        df = self.pareto_service.calculate_pareto(mode, category, days)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:H1')
        ws['A1'] = f'داده‌های خام - {mode} - {category}'
        ws['A1'].font = Font(name='B Nazanin', size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='37474F', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        headers = ['ردیف', 'کد کالا', 'نام کالا', 'مبلغ', 'درصد', 'مبلغ تجمعی', 'درصد تجمعی', 'کلاس']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='607D8B', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        if not df.empty:
            for r_idx, row in enumerate(df.itertuples(), 4):
                ws.cell(row=r_idx, column=1, value=row.row_num).border = thin_border
                ws.cell(row=r_idx, column=2, value=row.item_code).border = thin_border
                ws.cell(row=r_idx, column=3, value=row.item_name).border = thin_border
                
                amount_cell = ws.cell(row=r_idx, column=4, value=row.amount)
                amount_cell.border = thin_border
                amount_cell.number_format = '#,##0'
                
                pct_cell = ws.cell(row=r_idx, column=5, value=row.percentage / 100)
                pct_cell.border = thin_border
                pct_cell.number_format = '0.00%'
                
                cum_amount_cell = ws.cell(row=r_idx, column=6, value=row.cumulative_amount)
                cum_amount_cell.border = thin_border
                cum_amount_cell.number_format = '#,##0'
                
                cum_pct_cell = ws.cell(row=r_idx, column=7, value=row.cumulative_percentage / 100)
                cum_pct_cell.border = thin_border
                cum_pct_cell.number_format = '0.00%'
                
                abc_cell = ws.cell(row=r_idx, column=8, value=row.abc_class)
                abc_cell.border = thin_border
                abc_cell.alignment = Alignment(horizontal='center')
                
                # Color coding for ABC
                if row.abc_class == 'A':
                    abc_cell.fill = PatternFill(start_color='C8E6C9', fill_type='solid')
                    abc_cell.font = Font(bold=True, color='1B5E20')
                elif row.abc_class == 'B':
                    abc_cell.fill = PatternFill(start_color='FFF9C4', fill_type='solid')
                    abc_cell.font = Font(bold=True, color='F57F17')
                else:
                    abc_cell.fill = PatternFill(start_color='FFCCBC', fill_type='solid')
                    abc_cell.font = Font(color='BF360C')
                
                for col in range(1, 9):
                    ws.cell(row=r_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
        
        # Chart: Top 10 items by amount
        top_n = min(10, len(df))
        if top_n > 0:
            data_ref = Reference(ws, min_col=4, min_row=3, max_row=3 + top_n)
            cats_ref = Reference(ws, min_col=3, min_row=4, max_row=3 + top_n)
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.title = f"Top {top_n} اقلام بر اساس مبلغ"
            bar_chart.y_axis.title = 'مبلغ'
            bar_chart.x_axis.title = 'کالا'
            bar_chart.width = 18
            bar_chart.height = 10
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats_ref)
            ws.add_chart(bar_chart, "J4")
        
        for col, width in [('A', 8), ('B', 12), ('C', 20), ('D', 15), ('E', 10), ('F', 15), ('G', 12), ('H', 8)]:
            ws.column_dimensions[col].width = width
    
    def save_to_bytes(self, wb):
        """Save workbook to BytesIO for download"""
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
